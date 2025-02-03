import sys
from datetime import datetime
from dataclasses import dataclass
from openpyxl import load_workbook


# Data class to represent each procedure
@dataclass
class Procedure:
    number: str  # Number of the procedure, format: PROCXXX
    part: int  # Number of the part (If a long procedure is partionnated)
    title: str
    ignored: bool  # Check if the procedure has to be ignored.


# Class to manage the procedures
class ProcedureManager:
    def __init__(self, path: str) -> None:
        self.path = path
        self.procedures_list_file = "procedure_list.xlsx"
        self.history_file = "discussions_history.xlsx"
        self.procedures = []  
        self.__load_procedures()
        self.procedure_history = {procedure.number: None for procedure in self.procedures if not procedure.ignored} # Map to each procedure the last date it has been reviewed
        self.__generate_history()

    def __load_procedures(self) -> None:
        """Load all the procedures details from a file as a Procedure object and store them in self.procedures"""

        procedures_path = self.path + self.procedures_list_file
        
        wb = load_workbook(procedures_path)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            self.procedures.append(Procedure(*row))
        return
    
    
    def __generate_history(self) -> None:
        """Load the history file and get the last reviewed date for each procedure mapped in self.procedure_history"""
        
        history_path = self.path + self.history_file
        
        wb = load_workbook(history_path)
        sheet = wb.active

        for proc, date in sheet.iter_rows(min_row=2, values_only=True):
            if proc in self.procedure_history:
                self.procedure_history[proc] = max(self.procedure_history[proc], date) if self.procedure_history[proc] else date
        return
        

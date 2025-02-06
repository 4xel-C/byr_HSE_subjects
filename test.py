import unittest
from unittest.mock import patch, MagicMock
from datetime import datetime
from openpyxl import Workbook

from utils import select_quarter, select_generation_method, select_automatic_generation_menu, manual_selection_procedure_menu, ProcedureManager, Procedure

class TestMenuSelection(unittest.TestCase):
    
    # ----------------------------------------Quarter selection Test
    # test selection 1 for quarter selection
    @patch("builtins.input", side_effect=["1"])
    def test_select_quarter(self, mock_input):
        self.assertEqual(select_quarter(), 1)

    # test selection 5 for quarter selection
    @patch("builtins.input", side_effect=["5"])
    def test_select_quarter_exit(self, mock_input):
        with self.assertRaises(SystemExit):
            select_quarter()

    # test 2 bad selection, then a selection with a space for quarter selection
    @patch("builtins.input", side_effect=["abc", "7", "2  "])
    def test_select_quarter_invalid_then_valid(self, mock_input):
        self.assertEqual(select_quarter(), 2)

    # test selection 1 for quarter selection
    @patch("builtins.input", side_effect=["1"])
    def test_select_generation_method(self, mock_input):
        self.assertEqual(select_generation_method(1), "auto")

    # ----------------------------------------Method selection test
    @patch("builtins.input", side_effect=["2"])
    def test_select_generation_method_manual(self, mock_input):
        self.assertEqual(select_generation_method(1), "manual")

    @patch("builtins.input", side_effect=["sdqfkjh", "76", "3  "])
    def test_select_generation_method_invalid_then_valid(self, mock_input):
        self.assertEqual(select_generation_method(1), "back")

    @patch("builtins.input", side_effect=["4"])
    def test_select_generation_method_exit(self, mock_input):
        with self.assertRaises(SystemExit):
            select_generation_method(1)
            
    # -------------------------------------------------selection automatic generation test
    @patch("builtins.input", side_effect=["1"])
    def test_select1_automatic_generation_menu(self, mock_input):
        
        procedures = [
            Procedure(number="PROC001", part=1, title="Test title", ignored=False),
            Procedure(number="PROC002", part=1, title="Test title2", ignored=False),
            Procedure(number="PROC003", part=1, title="Test title3", ignored=False)
            ]
        
        months = ["January", "February", "March"]
        
        quarter = 1
        
        selection = select_automatic_generation_menu(procedures, quarter, months)
        self.assertEqual(selection, "0")
        
    @patch("builtins.input", side_effect=["3"])
    def test_select3_automatic_generation_menu(self, mock_input):
        
        procedures = [
            Procedure(number="PROC001", part=1, title="Test title", ignored=False),
            Procedure(number="PROC002", part=1, title="Test title2", ignored=False),
            Procedure(number="PROC003", part=1, title="Test title3", ignored=False)
            ]
        
        months = ["January", "February", "March"]
        
        quarter = 1
        
        selection = select_automatic_generation_menu(procedures, quarter, months)
        self.assertEqual(selection, "2")
    
    @patch("builtins.input", side_effect=["4"])
    def test_select_confirm_automatic_generation_menu(self, mock_input):
        
        procedures = [
            Procedure(number="PROC001", part=1, title="Test title", ignored=False),
            Procedure(number="PROC002", part=1, title="Test title2", ignored=False),
            Procedure(number="PROC003", part=1, title="Test title3", ignored=False)
            ]
        
        months = ["January", "February", "March"]
        
        quarter = 1
        
        selection = select_automatic_generation_menu(procedures, quarter, months)
        self.assertEqual(selection, "confirm")
        
    @patch("builtins.input", side_effect=["6"])
    def test_select_exit_automatic_generation_menu(self, mock_input):
        
        procedures = [
            Procedure(number="PROC001", part=1, title="Test title", ignored=False),
            Procedure(number="PROC002", part=1, title="Test title2", ignored=False),
            Procedure(number="PROC003", part=1, title="Test title3", ignored=False)
            ]
        
        months = ["January", "February", "March"]
        
        quarter = 1
        
        with self.assertRaises(SystemExit):
            select_automatic_generation_menu(procedures, quarter, months)
            
    # ------------------------------- Manual method menu
    
    @patch("builtins.input", side_effect=["1"])
    def test_select_manual_generation_menu(self, mock_input):
        
        procedures = [
            Procedure(number="PROC001", part=1, title="Test title", ignored=False),
            Procedure(number="PROC002", part=1, title="Test title2", ignored=False),
            Procedure(number="PROC003", part=1, title="Test title3", ignored=False)
            ]

        selection = manual_selection_procedure_menu(procedures)
        
        self.assertEqual(selection, "0")
            
class TestProcedure(unittest.TestCase):
    
    # Testing initializing procedure object
    def test_procedure_initialization(self):
        
        procedure = Procedure(number="PROC001", part=1, title="Test title", ignored=False)

        self.assertEqual(procedure.number, "PROC001")
        self.assertEqual(procedure.part, 1)
        self.assertEqual(procedure.title, "Test title")
        self.assertFalse(procedure.ignored)


class TestProcedureManager(unittest.TestCase):

    # setup method to create fake workbooks
    def setUp(self):
        
        # Create a fake work book procedure
        self.mock_workbook_proc = MagicMock()
        mock_sheet = self.mock_workbook_proc.active
        mock_sheet.iter_rows.return_value = [
            ("PROC001", 1, "Title 1", False),  
            ("PROC002", 2, "Title 2", False),  
            ("PROC003", None, "Title 3", False),  
            ("PROC004 ", None, "Title 3", False),  
            (" PROC005 ", None, "Title 3", False),  
        ]
        
        #  create a fake workbook history
        self.mock_workbook_history = MagicMock()
        mock_sheet = self.mock_workbook_history.active
        mock_sheet.iter_rows.return_value = [
            ("PROC001", datetime(2022, 6, 1, 0, 0)),  
            ("PROC002", datetime(2022, 7, 1, 0, 0)),
            ("PROC001 ", datetime(2022, 8, 1, 0, 0)),
            (("PROC005"), datetime(2022, 9, 1, 0, 0)),
            (("PROCXYZ"), datetime(2022, 12, 1, 0, 0)),
        ]
        

    # mock the load_workbook method 
    @patch('utils.procedures.load_workbook')
    def test_load_procedures(self, mock_load_workbook):
        
        # Two 'load_workbook' function are called when instanciating the ProcedureManager: 1 to load the procedures, the other, the history.
        mock_load_workbook.side_effect = [self.mock_workbook_proc, self.mock_workbook_history]
        

        # Initialize ProcedureManager
        manager = ProcedureManager(path='/fake/path/')

        # Check if procedure are correctly charged
        self.assertEqual(manager.procedures[0].number, "PROC001")
        self.assertEqual(manager.procedures[1].number, "PROC002")
        self.assertEqual(manager.procedures[3].number, "PROC004")
        self.assertEqual(manager.procedures[4].number, "PROC005")
        self.assertEqual(manager.procedures[2].title, "Title 3")
        self.assertIs(manager.procedures[2].part, None)
        self.assertEqual(len(manager.procedures), 5)
        self.assertEqual(manager.procedures[2].last_review, datetime.min)


    @patch('utils.procedures.load_workbook')
    def test_generate_history(self, mock_load_workbook):
        
        # Two 'load_workbook' function are called when instanciating the ProcedureManager: 1 to load the procedures, the other, the history.
        mock_load_workbook.side_effect = [self.mock_workbook_proc, self.mock_workbook_history]
        
        # Initialize ProcedureManager
        manager = ProcedureManager(path='/fake/path/')
        
        # Check if dictionnary has been correctly created
        self.assertEqual(manager.procedures[0].last_review, datetime(2022, 8, 1, 0, 0)) 
        self.assertEqual(manager.procedures[1].last_review, datetime(2022, 7, 1, 0, 0))
        self.assertEqual(manager.procedures[4].last_review, datetime(2022, 9, 1, 0, 0))
        self.assertEqual(manager.procedures[2].last_review, datetime.min),
        self.assertEqual(manager.procedures[4].last_review, datetime(2022, 9, 1, 0, 0))

    @patch('utils.procedures.load_workbook')
    def test_get_procedure_3(self, mock_load_workbook):
        
        # Two 'load_workbook' function are called when instanciating the ProcedureManager: 1 to load the procedures, the other, the history.
        mock_load_workbook.side_effect = [self.mock_workbook_proc, self.mock_workbook_history]
        
        # Initialize ProcedureManager
        manager = ProcedureManager(path='/fake/path/')

        selected = manager.get_procedures(3)

        self.assertEqual(len(selected), 3)

        # Check if there are sorted by data
        self.assertLessEqual(selected[0].last_review, selected[1].last_review)
        self.assertLessEqual(selected[1].last_review, selected[2].last_review)

        # Check the 3 older procedures
        self.assertEqual(selected[0].number, "PROC003")  
        self.assertEqual(selected[1].number, "PROC004")  
        self.assertEqual(selected[2].number, "PROC002")  

    @patch('utils.procedures.load_workbook')
    def test_get_procedure_2(self, mock_load_workbook):
        
        # Two 'load_workbook' function are called when instanciating the ProcedureManager: 1 to load the procedures, the other, the history.
        mock_load_workbook.side_effect = [self.mock_workbook_proc, self.mock_workbook_history]
        
        # Initialize ProcedureManager
        manager = ProcedureManager(path='/fake/path/')

        selected = manager.get_procedures(2)

        self.assertEqual(len(selected), 2)

        # Check if there are sorted by data
        self.assertLessEqual(selected[0].last_review, selected[1].last_review)

        # Check the 3 older procedures
        self.assertEqual(selected[0].number, "PROC003")  
        self.assertEqual(selected[1].number, "PROC004") 
        
    @patch('utils.procedures.load_workbook')
    def test_get_procedure_all(self, mock_load_workbook):
        
        # Two 'load_workbook' function are called when instanciating the ProcedureManager: 1 to load the procedures, the other, the history.
        mock_load_workbook.side_effect = [self.mock_workbook_proc, self.mock_workbook_history]
        
        # Initialize ProcedureManager
        manager = ProcedureManager(path='/fake/path/')

        selected = manager.get_procedures()

        self.assertEqual(len(selected), len(manager.procedures))

        # Check if there are sorted by data
        self.assertLessEqual(selected[0].last_review, selected[1].last_review)

        # Check the 3 older procedures
        self.assertEqual(selected[0].number, "PROC003")  
        self.assertEqual(selected[1].number, "PROC004") 



# run the tests
if __name__ == "__main__":
    unittest.main()
